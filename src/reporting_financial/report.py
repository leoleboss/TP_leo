from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for col_idx, column_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)


def format_sheet(ws):
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3


def create_report(
    df: pd.DataFrame,
    global_indicators: pd.DataFrame,
    client_type_indicators: pd.DataFrame,
    agent_indicators: pd.DataFrame,
    output_path: str = "output/reporting_financial.xlsx",
):
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    ws_clients = wb.create_sheet("Clients")
    ws_agents = wb.create_sheet("Agents")

    ws_dashboard["A1"] = "Reporting financier automatisé"
    ws_dashboard["A1"].font = Font(size=18, bold=True)
    ws_dashboard["A1"].alignment = Alignment(horizontal="center")

    write_dataframe(ws_dashboard, global_indicators, start_row=3, start_col=1)
    write_dataframe(ws_dashboard, client_type_indicators, start_row=3, start_col=4)

    write_dataframe(ws_clients, df, start_row=1, start_col=1)
    write_dataframe(ws_agents, agent_indicators, start_row=1, start_col=1)

    # Tableau de répartition des risques
    risk_counts = df["risk_level"].value_counts().reset_index()
    risk_counts.columns = ["Niveau de risque", "Nombre"]
    write_dataframe(ws_dashboard, risk_counts, start_row=12, start_col=1)

    # Graphique barres : clients par type
    bar_chart = BarChart()
    bar_chart.title = "Nombre de clients par type"
    bar_chart.y_axis.title = "Nombre de clients"
    bar_chart.x_axis.title = "Type client"

    data = Reference(ws_dashboard, min_col=5, min_row=3, max_row=5)
    categories = Reference(ws_dashboard, min_col=4, min_row=4, max_row=5)

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)

    ws_dashboard.add_chart(bar_chart, "G3")

    # Graphique camembert : niveaux de risque
    pie_chart = PieChart()
    pie_chart.title = "Répartition par niveau de risque"

    pie_data = Reference(ws_dashboard, min_col=2, min_row=12, max_row=15)
    pie_labels = Reference(ws_dashboard, min_col=1, min_row=13, max_row=15)

    pie_chart.add_data(pie_data, titles_from_data=True)
    pie_chart.set_categories(pie_labels)

    ws_dashboard.add_chart(pie_chart, "G18")

    for ws in wb.worksheets:
        format_sheet(ws)

    wb.save(output_file)

    print(f"Rapport généré avec succès : {output_file}")
    